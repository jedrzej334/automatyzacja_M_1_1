import tkinter as tk
from tkinter import simpledialog, messagebox
from openpyxl import load_workbook
import clr
import math

# Dodanie odniesień do bibliotek APx500
clr.AddReference("System.Drawing")
clr.AddReference("System.Windows.Forms")
clr.AddReference(r"C:\\Program Files\\Audio Precision\\APx500 8.1\\API\\AudioPrecision.API2.dll")
clr.AddReference(r"C:\\Program Files\\Audio Precision\\APx500 8.1\\API\\AudioPrecision.API.dll")

from AudioPrecision.API import APx500_Application, APxOperatingMode, OutputChannelIndex
from System.IO import Directory, Path

# Inicjalizacja aplikacji APx500
APx = APx500_Application(APxOperatingMode.BenchMode, True)

class AudioInterfaceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Interfejs Audio")
        self.root.geometry("500x700")

        ##### Zmienne początkowe #####
        self.initial_voltage = 0.0001  # Początkowe napięcie (w Voltach) - ustawione na 100 uVrms
        self.level_V = self.initial_voltage  # Aktualne napięcie
        self.value_dB = 0.0  # Początkowy poziom dB
        self.freqList = [63, 125, 250, 500, 1000, 2000, 4000, 8000, 16000]  # Częstotliwości
        self.freq_var = tk.IntVar(value=0)  # Domyślna częstotliwość: 63 Hz

        # Inicjalizacja listy do przechowywania wartości dB
        self.values_list = []

        ##### Tworzenie GUI #####
        # Nagłówek
        self.label = tk.Label(self.root, text="Wybierz częstotliwość", font=("Arial", 14))
        self.label.place(x=150, y=20)

        # Suwak do wyboru częstotliwości
        self.freq_label = tk.Label(self.root, text="Częstotliwość (Hz):", font=("Arial", 12))
        self.freq_label.place(x=50, y=70)

        self.frequency_slider = tk.Scale(self.root, from_=0, to=len(self.freqList) - 1, orient=tk.HORIZONTAL,
                                         label="Częstotliwość (Hz)", variable=self.freq_var, tickinterval=1)
        self.frequency_slider.place(x=50, y=110, width=400)

        self.freq_value_label = tk.Label(self.root, text=f"Wybrana częstotliwość: {self.freqList[self.freq_var.get()]} Hz", font=("Arial", 12))
        self.freq_value_label.place(x=150, y=150)

        self.frequency_slider.bind("<Motion>", self.update_frequency_label)

        # Przycisk do zmiany częstotliwości na 500 Hz
        self.change_freq_button = tk.Button(self.root, text="Ustaw 500 Hz", command=lambda: self.set_frequency(500))
        self.change_freq_button.place(x=150, y=200)

        # Poziom dB
        self.level_label = tk.Label(self.root, text="Poziom (dB):", font=("Arial", 12))
        self.level_label.place(x=50, y=250)

        self.level_value_label = tk.Label(self.root, text=f"{self.value_dB:.1f} dB", font=("Arial", 12))
        self.level_value_label.place(x=150, y=290)

        # Przyciski do zmiany poziomu dB
        self.increase_01_button = tk.Button(self.root, text="Zwiększ o 0.1 dB", command=lambda: self.change_dB(0.1))
        self.increase_01_button.place(x=50, y=330)

        self.decrease_01_button = tk.Button(self.root, text="Zmniejsz o 0.1 dB", command=lambda: self.change_dB(-0.1))
        self.decrease_01_button.place(x=200, y=330)

        # Przycisk do rozpoczęcia pomiarów
        self.start_button = tk.Button(self.root, text="Rozpocznij Pomiar", command=self.start_measurement)
        self.start_button.place(x=150, y=380)

    def setGeneratorParams(self, level_V):
        # Jeśli napięcie jest większe niż 2V, ograniczamy je do 2V
        if level_V > 2.0:
            level_V = 2.0  # Ustawienie maksymalnego napięcia na 2 Vrms
        APx.BenchMode.Generator.Levels.SetValue(OutputChannelIndex.Ch1, level_V)
        APx.BenchMode.Generator.On = True

    def set_frequency(self, freq):
        if freq in self.freqList:
            self.freq_var.set(self.freqList.index(freq))
            self.freq_value_label.config(text=f"Wybrana częstotliwość: {freq} Hz")
            APx.BenchMode.Generator.Frequency.Value = freq

    def change_dB(self, step):
        self.value_dB += step
        if self.value_dB < 0:
            self.value_dB = 0
        self.level_value_label.config(text=f"{self.value_dB:.1f} dB")
        self.setGeneratorParams(self.level_V)

    def update_frequency_label(self, event):
        self.freq_value_label.config(text=f"Wybrana częstotliwość: {self.freqList[self.freq_var.get()]} Hz")
        currentFreq = self.freqList[self.freq_var.get()]
        APx.BenchMode.Generator.Frequency.Value = currentFreq

    def start_measurement(self):
        # Ustawienie początkowego napięcia
        user_input_voltage = simpledialog.askfloat("Ustaw początkowe napięcie", "Podaj początkowe napięcie (V) dla 1000 Hz:", minvalue=0.001, maxvalue=2.0)
        
        if user_input_voltage is not None:
            # Weryfikacja wprowadzonego napięcia
            is_correct = messagebox.askyesno("Weryfikacja napięcia", f"Czy napięcie {user_input_voltage} V jest poprawne?")
            if not is_correct:
                return  # Jeśli użytkownik wybierze "Nie", nie kontynuujemy

            self.initial_voltage = user_input_voltage
            self.level_V = self.initial_voltage
            self.setGeneratorParams(self.level_V)

            # Dodanie komunikatu, że napięcie zostało ustawione
            messagebox.showinfo("Ustawienie napięcia", f"Napięcie ustawione na {self.level_V} V")

        # Zmiana częstotliwości na 500 Hz i zwiększenie napięcia o 3.2 dB
        self.set_frequency(500)
        self.change_voltage_dB(3.2)

        # Po 500 Hz przejdź na 250 Hz i zwiększ napięcie o 8.6 dB
        self.set_frequency(250)
        self.change_voltage_dB(8.6)

        # Pytanie o wartość dB i zapisanie jej do pliku Excel
        user_input_dB = simpledialog.askfloat("Wprowadź wartość", "Podaj wartość dB z miernika:")
        if user_input_dB is not None:
            self.values_list.append(user_input_dB)
            print(self.values_list)

            # Zapisz dane do pliku Excel
            try:
                wb = load_workbook('C:/Users/akust/Desktop/Automatyzacja_Jędrzej/automatyzacja_test.xlsx')
                ws = wb.active
                for i, value in enumerate(self.values_list, start=1):
                    ws[f'P{i}'] = value
                wb.save('C:/Users/akust/Desktop/Automatyzacja_Jędrzej/automatyzacja_test.xlsx')
                print("Dane zapisane do pliku Excel.")
            except FileNotFoundError:
                print("Nie znaleziono pliku Excel. Upewnij się, że ścieżka jest poprawna.")

    def change_voltage_dB(self, dB_increase):
        # Przeliczanie napięcia na podstawie zmiany dB
        # Przeliczanie napięcia w zależności od zmiany dB (używając wzoru: V2 = V1 * 10^(dB/20))
        new_voltage = self.level_V * 10 ** (dB_increase / 20)

        # Sprawdzenie, czy napięcie nie przekracza 2 Vrms
        if new_voltage > 2.0:
            new_voltage = 2.0  # Ustawienie maksymalnego napięcia na 2 Vrms

        self.level_V = new_voltage
        self.setGeneratorParams(self.level_V)


# Tworzenie okna aplikacji
root = tk.Tk()
app = AudioInterfaceApp(root)
root.mainloop()
