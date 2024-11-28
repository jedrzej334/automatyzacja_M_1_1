import tkinter as tk
from tkinter import simpledialog  # Importujemy simpledialog do tworzenia okna dialogowego
from openpyxl import load_workbook
import sys
import clr
import time

# Dodanie referencji do systemowych bibliotek
clr.AddReference("System.Drawing")
clr.AddReference("System.Windows.Forms")

# Dodanie referencji do bibliotek APx500
clr.AddReference(r"C:\\Program Files\\Audio Precision\\APx500 8.1\\API\\AudioPrecision.API2.dll")
clr.AddReference(r"C:\\Program Files\\Audio Precision\\APx500 8.1\\API\\AudioPrecision.API.dll")

from AudioPrecision.API import *
from System.Drawing import Point
from System.Windows.Forms import Application, Button, Form, Label
from System.IO import Directory, Path

# Inicjalizacja APx500 w trybie BenchMode
APx = APx500_Application(APxOperatingMode.BenchMode, True)


class AudioInterfaceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Interfejs Audio")
        self.root.geometry("500x700")

        ##### Zmienne początkowe ######
        self.value_dB = 0.0  # Początkowy poziom dB
        self.freqList = [63, 125, 250, 500, 1000, 2000, 4000, 8000, 16000]  # Częstotliwości
        self.freq_var = tk.IntVar(value=0)  # Domyślna częstotliwość: 63 Hz
        self.values_list = []  # Lista do przechowywania wyników

        # Tworzenie GUI
        self.label = tk.Label(self.root, text="Wybierz częstotliwość", font=("Arial", 14))
        self.label.place(x=150, y=20)

        # Suwak do wyboru częstotliwości
        self.freq_label = tk.Label(self.root, text="Częstotliwość (Hz):", font=("Arial", 12))
        self.freq_label.place(x=50, y=70)

        self.frequency_slider = tk.Scale(self.root, from_=0, to=len(self.freqList) - 1, orient=tk.HORIZONTAL,
                                          label="Częstotliwość (Hz)", variable=self.freq_var,
                                          tickinterval=1)
        self.frequency_slider.place(x=50, y=110, width=400)

        self.freq_value_label = tk.Label(self.root, text=f"Wybrana częstotliwość: {self.freqList[self.freq_var.get()]} Hz",
                                         font=("Arial", 12))
        self.freq_value_label.place(x=150, y=150)

        self.frequency_slider.bind("<Motion>", self.update_frequency_label)

        # Okno do wyświetlania poziomu dB
        self.level_label = tk.Label(self.root, text="Poziom (dB):", font=("Arial", 12))
        self.level_label.place(x=50, y=200)

        self.level_value_label = tk.Label(self.root, text=f"{self.value_dB} dB", font=("Arial", 12))
        self.level_value_label.place(x=150, y=240)

        # Przyciski do zmiany poziomu dB
        self.increase_01_button = tk.Button(self.root, text="Zwiększ o 0.1 dB", command=lambda: self.change_dB(0.1))
        self.increase_01_button.place(x=50, y=280)

        self.decrease_01_button = tk.Button(self.root, text="Zmniejsz o 0.1 dB", command=lambda: self.change_dB(-0.1))
        self.decrease_01_button.place(x=200, y=280)

        # Nowy przycisk, który zmienia dB, częstotliwość i uruchamia okno dialogowe
        self.increase_3_2_button = tk.Button(self.root, text="+3,2", command=self.ask_for_value_and_change)
        self.increase_3_2_button.place(x=50, y=460)

    def setGeneratorParams(self, level_dB):
        """
        Ustaw poziom generatora w oparciu o podany poziom w dB.
        """
        try:
            # Konwersja z dB na napięcie RMS w Voltach
            level_V = 10 ** (level_dB / 20.0)  # dBV -> Voltage
            print(f"Ustawiam poziom generatora na: {level_V:.3f} V RMS")

            # Ustawianie poziomu generatora na kanał 1
            APx.BenchMode.Generator.Levels.SetValue(OutputChannelIndex.Ch1, level_V)
            APx.BenchMode.Generator.On = True
        except Exception as e:
            print(f"Błąd podczas ustawiania generatora: {e}")

    def ask_for_value_and_change(self):
        self.change_dB(3.2)  # Zwiększ poziom dB o 3.2
        self.set_frequency(500)  # Ustaw częstotliwość na 500 Hz

        # Okno dialogowe z pytaniem o liczbę
        user_input = simpledialog.askfloat("Wprowadź wartość", "Podaj wartość dB z miernika:")

        # Jeśli użytkownik podał liczbę (a nie anulował dialogu)
        if user_input is not None:
            self.values_list.append(user_input)  # Zapisz wartość do listy
            print(self.values_list)  # Wydrukuj zawartość listy dla debugowania

            # Otwórz i zapisz dane do pliku Excel
            try:
                wb = load_workbook('C:/Users/akust/Desktop/Automatyzacja_Jędrzej/automatyzacja_test.xlsx')  # Wczytaj plik Excel
                ws = wb.active
                for i, value in enumerate(self.values_list, start=1):  # Zapisz wartości do kolumny P
                    ws[f'P{i}'] = value
                wb.save('C:/Users/akust/Desktop/Automatyzacja_Jędrzej/automatyzacja_test.xlsx')  # Zapisz zmiany
                print("Dane zapisane do pliku Excel.")
            except FileNotFoundError:
                print("Nie znaleziono pliku Excel. Upewnij się, że ścieżka jest poprawna.")

    def change_dB(self, step):
        """
        Zmienia poziom dB i aktualizuje generator.
        """
        self.value_dB += step
        if self.value_dB < 0:
            self.value_dB = 0
        self.level_value_label.config(text=f"{self.value_dB:.1f} dB")

        # Aktualizacja generatora
        self.setGeneratorParams(self.value_dB)

    def set_frequency(self, freq):
        if freq in self.freqList:
            self.freq_var.set(self.freqList.index(freq))
            self.freq_value_label.config(text=f"Wybrana częstotliwość: {freq} Hz")

    def update_frequency_label(self, event):
        self.freq_value_label.config(text=f"Wybrana częstotliwość: {self.freqList[self.freq_var.get()]} Hz")
        currentFreq = self.freqList[self.freq_var.get()]
        APx.BenchMode.Generator.Frequency.Value = currentFreq


# Główna część aplikacji
if __name__ == "__main__":
    root = tk.Tk()
    app = AudioInterfaceApp(root)
    root.mainloop()

    # Wyłącz generator przy zamknięciu aplikacji
    APx.BenchMode.Generator.On = False
